"""

from firebase import firebase


db = firebase.FirebaseApplication(
    'https://sensorprueba-4924b-default-rtdb.firebaseio.com/')
"""

import pyrebase as bs
import json
from openpyxl import Workbook
import openpyxl


config = {
    "apiKey": "AIzaSyAg3Kb0scSwfw8cBSiCioIPEwmJG4qpcHw",
  "authDomain": "sensorprueba-4924b.firebaseapp.com",
  "databaseURL": "https://sensorprueba-4924b-default-rtdb.firebaseio.com",
  "projectId": "sensorprueba-4924b",
  "storageBucket": "sensorprueba-4924b.appspot.com",
  "messagingSenderId": "44066862233",
  "appId": "1:44066862233:web:d0bad738340d4e039af408"
}


firebase = bs.initialize_app(config)
db = firebase.database()






def nuevas_medidas(alturaCilindro, diametro, alturaCono,capacidad, usuario):
    
    try:
        iden = leer_identificador()


        key = ""
        lista_silos = db.child("silo").child(usuario).get()
        for silos in lista_silos.each():
            if(silos.val()['id']== iden):
                key=silos.key()

        db.child("silo").child(usuario).child(key).update({"alturaCilindro":alturaCilindro})
        db.child("silo").child(usuario).child(key).update({"capacidad":capacidad})
        db.child("silo").child(usuario).child(key).update({"alturaCono":alturaCono})
        db.child("silo").child(usuario).child(key).update({"diametro":diametro})
        return 0
    except:
        return -1


def leer_identificador():
    book = openpyxl.load_workbook('data/silo.xlsx')
    libro = book.active

    usuario = libro['A1']
    return usuario.value


def listar_datos(usuario):

    lista_de_tuplas = []
    users = db.child("silo/"+usuario).get()
    for dato in users.each():
        objeto = dato.val()
        lista_tupla = []
        objeto_string = json.dumps(objeto)
        objeto_silo = json.loads(objeto_string)
        lista_tupla.append(objeto_silo['id'])
        lista_tupla.append(objeto_silo['fecha'])
        lista_tupla.append(objeto_silo['inventario'])
        lista_tupla.append(objeto_silo['ocupacion'])
        lista_tupla.append(objeto_silo['estado'])
        lista_tupla.append(objeto_silo['alturaCilindro'])
        lista_tupla.append(objeto_silo['diametro'])
        lista_tupla.append(objeto_silo['alturaCono'])
        lista_tupla.append(objeto_silo['capacidad'])
        lista_de_tuplas.append(lista_tupla)
    return lista_de_tuplas


def buscar_silo(iden, usuario):

    try:
        datos = listar_datos(usuario)
        
        lista_silo = []
        i = 0
        while i < len(datos):
            if datos[i][0] == iden:
                lista_silo.append(datos[i])
                break
            else:
                i+=1
        
        print(lista_silo)
        hoja = Workbook()
        sheet = hoja.active
        sheet['A1'] = lista_silo[0][0]      
        hoja.save('data/silo.xlsx')
        return 0
    except:
        return -1
      
   





def usuario_logueado_excel(usuario, password):


    try:
        hoja = Workbook()
        sheet = hoja.active
        name = obtener_usuario(usuario)
        sheet['A1'] = name
        sheet['B1'] = password        

        hoja.save('data/usuario.xlsx')
        return 0
    except:
        return -1






def generar_excel(usuario):

    try:
        hoja = Workbook()
        sheet = hoja.active
        encabezado = ["id_silo", "fecha", "inventario",
                        "ocupación", "estado", "alturaCilindro", "diametro", "alturaCono", "capacidad"]

        letras = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
        for i in range(len(encabezado)):
            sheet[letras[i]+'1'] = encabezado[i]

        datas = listar_datos(usuario)
        i = 0
        while i < len(datas):
            sheet[f'A{i+2}'] = datas[i][0]
            sheet[f'B{i+2}'] = datas[i][1]
            sheet[f'C{i+2}'] = datas[i][2]
            sheet[f'D{i+2}'] = datas[i][3]
            sheet[f'E{i+2}'] = datas[i][4]
            sheet[f'F{i+2}'] = datas[i][5]
            sheet[f'G{i+2}'] = datas[i][6]
            sheet[f'H{i+2}'] = datas[i][7]
            sheet[f'I{i+2}'] = datas[i][8]
            i += 1

        hoja.save('data/datos.xlsx')
        return 0
    except:
        return -1


def usuario_logueado_excel(usuario, password):


    try:
        hoja = Workbook()
        sheet = hoja.active
        name = obtener_usuario(usuario)
        sheet['A1'] = name
        sheet['B1'] = password        

        hoja.save('data/usuario.xlsx')
        return 0
    except:
        return -1


def obtener_usuario(name):
    
    
    lista_nombre = list(name)

    i = 0
    usuario = ""
    while i < len(lista_nombre):
        if lista_nombre[i] != "@":
            usuario+=lista_nombre[i]
            i+=1
        else:
            break
    return usuario



def leer_usuario():

    book = openpyxl.load_workbook('data/usuario.xlsx')
    libro = book.active

    usuario = libro['A1']
    return usuario.value